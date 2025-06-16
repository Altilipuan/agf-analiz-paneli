# ------------------- AGF TAKİP VE ANALİZ SİSTEMİ -------------------

from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
import time

# ------------------- GİRİŞ -------------------
agf_url = input("AGF verilerini çekeceğimiz TJK linkini girin: ").strip()
saat_girdisi = input("Veri çekim saatlerini girin (14:00,15:15,16:30): ").strip()

planlanan_cekimler = []
for s in [x.strip() for x in saat_girdisi.split(",") if x.strip()]:
    try:
        dt = datetime.strptime(s, "%H:%M")
        planlanan_cekimler.append(dt.strftime("%H:%M"))
    except:
        print(f"⚠️ Geçersiz saat formatı atlandı: {s}")

agf_data_dict = {}
output_file = "agf_zaman_serisi_ve_analiz.xlsx"

# ------------------- SÜRPRİZ FONKSİYONU (GÜNCELLENDİ) -------------------
def belirle_surpriz_tipi(row, saatler):
    try:
        agf_values = row[1:-1].dropna().astype(float)
        if len(agf_values) < 3:
            return ""

        ilk_agf = agf_values.iloc[0]
        son_agf = agf_values.iloc[-1]
        fark_ilk_son = son_agf - ilk_agf

        if son_agf < 10 and fark_ilk_son >= 1.3:
            return "SÜRPRİZ"

        if len(saatler) >= 2:
            son1 = row[saatler[-1]]
            son2 = row[saatler[-2]]
            if pd.notna(son1) and pd.notna(son2):
                fark_son_dk = float(son1) - float(son2)
                if fark_son_dk >= 0.3 and float(son1) < 10:
                    return "Son DK Sürpriz"
    except:
        return ""
    return ""

# ------------------- VERİ ÇEK -------------------
def fetch_agf():
    timestamp = datetime.now().strftime("%H:%M")
    print(f"\n[{timestamp}] AGF verisi çekiliyor...")
    try:
        response = requests.get(agf_url)
        soup = BeautifulSoup(response.content, "html.parser")

        for ayak in range(1, 7):
            table = soup.find("table", {"id": f"GridView{ayak}"})
            if not table:
                continue
            rows = table.find_all("tr")[1:]
            current_data = []
            for row in rows:
                cells = row.find_all("td")
                if len(cells) >= 2:
                    cell_text = cells[1].text.strip()
                    if "(" in cell_text and "%" in cell_text:
                        at_no = cell_text.split("(")[0].strip()
                        agf_percent = cell_text.split("%")[-1].replace(")", "").replace(",", ".")
                        current_data.append((at_no, float(agf_percent)))

            df = pd.DataFrame(current_data, columns=["At", timestamp])
            if ayak not in agf_data_dict or agf_data_dict[ayak].empty:
                agf_data_dict[ayak] = df
            else:
                agf_data_dict[ayak] = pd.merge(agf_data_dict[ayak], df, on="At", how="outer")

        analyze_agf(timestamp)

        with pd.ExcelWriter(output_file, engine="openpyxl", mode="w") as writer:
            for ayak_no, df in agf_data_dict.items():
                df.to_excel(writer, sheet_name=f"Ayak_{ayak_no}", index=False)

        create_analysis_sheet()
        print(f"✅ {timestamp} - Veriler işlendi ve kayıt edildi.")
    except Exception as e:
        print("⚠️ Veri çekme hatası:", e)

# ------------------- ANALİZ -------------------
def analyze_agf(current_time):
    for ayak, df in agf_data_dict.items():
        if df.shape[1] < 3:
            continue
        saatler = df.columns[1:].tolist()
        last_col = df.columns[-1]
        prev_col = df.columns[-2]
        df["Δ AGF"] = df[last_col] - df[prev_col]
        df["Toplam Δ"] = df[last_col] - df[df.columns[1]]
        df["Volatilite"] = df[df.columns[1:-1]].std(axis=1)
        df["Trend Skoru"] = df[df.columns[1:-1]].diff(axis=1).apply(lambda x: sum([1 if v > 0 else -1 if v < 0 else 0 for v in x.dropna()]), axis=1)
        df["Sürpriz Tipi"] = df.apply(lambda row: belirle_surpriz_tipi(row, saatler), axis=1)

# ------------------- ANALİZ SAYFASI -------------------
        from openpyxl.styles import Color
        from openpyxl import load_workbook

        wb = load_workbook(output_file)
        if "Analiz" in wb.sheetnames:
            del wb["Analiz"]
        ws = wb.create_sheet("Analiz")
        ws.sheet_view.showGridLines = False

        fill_blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        fill_yellow = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        fill_red = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
        fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        bold = Font(bold=True)
        green_font = Font(bold=True, color="006100")
        center = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        row_offset = 1
        for ayak in range(1, 7):
            if ayak not in agf_data_dict or agf_data_dict[ayak].shape[1] < 3:
                continue

            df = agf_data_dict[ayak]
            saatler = df.columns[1:].tolist()
            last_col = df.columns[-1]
            prev_col = df.columns[-2]
            df["Δ AGF"] = df[last_col] - df[prev_col]
            df["Toplam Δ"] = df[last_col] - df[df.columns[1]]
            df["Volatilite"] = df[df.columns[1:-1]].std(axis=1)
            df["Trend Skoru"] = df[df.columns[1:-1]].diff(axis=1).apply(lambda x: sum([1 if v > 0 else -1 if v < 0 else 0 for v in x.dropna()]), axis=1)
            df["Sürpriz Tipi"] = df.apply(lambda row: belirle_surpriz_tipi(row, saatler), axis=1)

            df = df.sort_values(by=["Trend Skoru", "Toplam Δ", "Volatilite"], ascending=[False, False, False])

            max_trend = df["Trend Skoru"].max()
            max_delta = df["Toplam Δ"].max()
            max_vol = df["Volatilite"].max()

            surpriz = df[df["Sürpriz Tipi"] != ""]
            max_surpriz_agf = (surpriz[last_col] - surpriz[df.columns[1]]).max() if not surpriz.empty else None

            ws.merge_cells(start_row=row_offset, start_column=1, end_row=row_offset, end_column=5)
            cell = ws.cell(row=row_offset, column=1, value=f"{ayak}. AYAK")
            cell.font = Font(bold=True, size=12)
            cell.alignment = center
            cell.fill = fill_blue
            cell.border = border

            headers = ["At", "Sürekli Artış Göstermiş Atlar", "Toplam AGF Değişimi", "Sabit Çok Değişmeyen Atlar"]
            for col_index, title in enumerate(headers, start=1):
                cell = ws.cell(row=row_offset + 1, column=col_index, value=title)
                cell.font = bold
                cell.alignment = center
                cell.fill = fill_blue
                cell.border = border

            for i, row in df.iterrows():
                r = row_offset + 2 + i
                for col, key in zip([1, 2, 3, 4], ["At", "Trend Skoru", "Toplam Δ", "Volatilite"]):
                    cell = ws.cell(row=r, column=col, value=row[key])
                    cell.alignment = center
                    cell.border = border
                    if (key == "Trend Skoru" and row[key] == max_trend) or \
                       (key == "Toplam Δ" and row[key] == max_delta) or \
                       (key == "Volatilite" and row[key] == max_vol):
                        cell.fill = fill_green
                        ws.cell(row=r, column=1).fill = fill_green
                        ws.cell(row=r, column=1).font = green_font

            msg_row = row_offset + 2 + len(df) + 1
            for _, row in surpriz.iterrows():
                artis = row[last_col] - row[df.columns[1]]
                tip = row["Sürpriz Tipi"]
                if tip == "SÜRPRİZ":
                    text = f"⚠️ SÜRPRİZ: {row['At']} (%+{artis:.1f})"
                    fill = fill_red
                else:
                    text = f"⏱️ Son DK Sürpriz: {row['At']} (%+{artis:.1f})"
                    fill = fill_yellow
                cell = ws.cell(row=msg_row, column=1, value=text)
                cell.fill = fill
                cell.font = bold
                cell.alignment = center
                if artis == max_surpriz_agf:
                    cell.fill = fill_green
                    cell.font = green_font
                msg_row += 1

            for _, row in df[df["Δ AGF"] > 2].iterrows():
                text = f"Sinyal: {row['At']} (+{row['Δ AGF']:.1f})"
                cell = ws.cell(row=msg_row, column=1, value=text)
                cell.fill = fill_yellow
                cell.font = bold
                cell.alignment = center
                msg_row += 1

            row_offset = msg_row + 2

        wb.save(output_file)
        print("🧩 Analiz sayfası estetik biçimde güncellendi.")

# ------------------- BAŞLANGIÇ -------------------
print("\n🚀 AGF takip sistemi başlatıldı...")
print(f"⏱️ Planlanan çekim saatleri: {planlanan_cekimler}\n")

while True:
    now = datetime.now().strftime("%H:%M")
    if now in planlanan_cekimler:
        fetch_agf()
        planlanan_cekimler.remove(now)
    if not planlanan_cekimler:
        print("✅ Tüm veri çekimleri başarıyla tamamlandı.")
        break
    time.sleep(30)
